from openpyxl import load_workbook

##################### FUNÇÕES

def escrevecsv(string_completa):
    f2 = open('excel.csv','a',encoding="utf8")
    f2.write(string_completa+"\n")
    f2.close()


def salvaArquivo(file):
    book.save(file)
    

def retira_acentuacoes_erradas(string):
    return string.replace("Ã”","Ô").replace("Ã“","Ó").replace("ÃŠ","Ê").replace("Ã‡","Ç").replace("Ãœ","U").replace("Ã‰","É").replace("Ãƒ","Ã").replace("Ãš","Ú")

def retira_espacamentos(string):
    return string.strip().replace('A VENIDA','AVENIDA').replace('JA RDIM','JARDIM').replace('JAR DIM','JARDIM')

def le_arquivo(letra):
    f = open("ENDEREÇOS CONTRIBUINTES DE GUARARAPES.txt","r",encoding="utf8")
    f1 = f.readlines()

    contador=0 # contador da lista f1
    proxima_linha_rua=0 # variavel para trablhar com a proxima linha na mesma string
    proxima_linha_bairro=0 # somente bairro
    string_completa=""

    contador1=0

    
    for linha in f1:
        if 'Cadastro:' in linha:
            continue

        if proxima_linha_rua == 1:
            ## Verifica se é rua + bairro OU somente rua

            if 'Cadastro:' in linha or linha.startswith("1"):
                continue
            if 'RUA' in linha:            
                if string_completa != "":
                        
                    rua=retira_acentuacoes_erradas(linha[0:linha.find(',')+7])
                    if linha[linha.find(',')+7:-1] != "":
                        # TEM BAIRRO
                        string_completa+='|'+rua+'|'+ linha[linha.find(',')+7:-1]
                            
                        escrevecsv(string_completa)
                        proxima_linha_rua=0
                        continue
            proxima_linha_rua=0
                    
        if proxima_linha_bairro == 1:
            string_completa+=retira_acentuacoes_erradas(linha.strip('\n'))
            # chamar função para dar append ao arquivo .csv
            escrevecsv(string_completa)
            proxima_linha_bairro=0
            continue
            
        if linha.startswith(letra) and linha.split(' ')[0] != 'RUA' and 'RUA' in linha or linha.startswith(letra) and 'AVENIDA' in linha:
            contador1+=1
            if 'RUA' in linha:
                    nome = retira_acentuacoes_erradas(linha[0:linha.find('RUA')])
                    delimitador_virgula=linha.find(',')+7
                    rua=retira_acentuacoes_erradas(linha[linha.find('RUA'):delimitador_virgula])
                    
                    # verifica se após a "," existe algo , ou seja, o BAIRRO;
                    if linha[delimitador_virgula:-1] != "":
                        # Bairro existe
                         bairro=retira_acentuacoes_erradas((linha[delimitador_virgula:-1]))
                         string_completa=nome+'|'+rua+'|'+bairro
                         escrevecsv(string_completa)
                    else:
                        # Sem bairro
                        # Preciso pegar a linha de baixo
                        string_completa=nome+'|'+rua+'|'
                        proxima_linha_bairro+=1
                            
            elif 'AVENIDA' in linha:
                nome = retira_acentuacoes_erradas(linha[0:linha.find('AVENIDA')])
                delimitador_virgula=linha.find(',')+7
                avenida= retira_acentuacoes_erradas(linha[linha.find('AVENIDA'):delimitador_virgula])
                # verifica se após a "," existe algo , ou seja, o BAIRRO;
                if linha[delimitador_virgula:-1] != "":
                # Bairro existe
                    bairro=retira_acentuacoes_erradas(linha[delimitador_virgula:-1])
                    string_completa=nome+'|'+avenida+'|'+bairro
                    escrevecsv(string_completa)
                else:
                            # Sem bairro
                            # Preciso pegar a linha de baixo
                    string_completa=nome+'|'+avenida+'|'
                    proxima_linha_bairro+=1
                            
        else:
            if linha.startswith(letra) and linha.split(' ')[0] != "RUA":
                contador1+=1
                ## Se a rua estiver na linha de baixo
                string_completa=linha.strip('\n') # Somente o nome
                proxima_linha_rua+=1
                continue
        

################### FIM FUNÇÕES







#################### MAIN CODE



while 1:
    letra=input("~~> Qual letra deve-se procurar ?(0) para verificar os ceps ")
    if letra == "0":
        break
    le_arquivo(letra)


print("Modifique a extensão de excel.csv para excel.xlsx.")
input("ENTER PARA PROSSEGUIR")

file='excel.xlsx'
book=load_workbook(file)
sheet=book[book.sheetnames[0]]

print("[+] Arrumando os cep's...")

for i in range(1,sheet.max_row):
    string="C"+str(i)
    valorA1=""
    
    try:
        valorA1=retira_espacamentos(sheet[string].value.strip())
    except:
        continue
    
    if 'GUARARAPES -  SP 16700000GUARARAPES - SP 16700000' in valorA1:
        novo_valor=valorA1.replace("GUARARAPES -  SP 16700000GUARARAPES - SP 16700000","GUARARAPES - SP 16700000")
        sheet[string]=novo_valor
        salvaArquivo(file)
    elif 'GUARARAPES' in valorA1 or 'GUARA RAPES' in valorA1:
        print(valorA1)
        continue
    else:
        #'GUARARAPES' not in valorA1 or 'GUARA RAPES' not in valorA1:
        # nao tem guararapes. Adicione
        novo_valor=valorA1+ " GUARARAPES - SP 16700000"
        sheet[string]=novo_valor
        salvaArquivo(file)
    
    print(valorA1 + "---> "+novo_valor)



